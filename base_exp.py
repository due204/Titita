from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import filedialog
from tkinter import messagebox
from base import BaseDatos
from datetime import datetime


# Esto es para guardar los datos en la DB
def importacion():
    importante = messagebox.askyesno(
        message="""Esta opcion sobre escribira su
base de datos actual para restaurar una copia de seguridad, asegurese de no
tener datos importantes en su base de datos.
\n¿Desea continuar?""",
        title="Importante...",
        # parent="destruct_button",
    )
    if importante:
        guard = BaseDatos()
        resultado = BaseDatos.select()
        # Ruta al archivo xlsx
        ruta = filedialog.askopenfilename(
            title="Importar Base",
            filetypes=(
                ("Archivo de exel", "*.xlsx"),
                ("Archivo de exel", "*.xlsm"),
                ("Archivo de exel", "*.xltx"),
                ("Archivo de exel", "*.xltm"),
                ("Todos los archivos", "*.*"),
            ),
        )
        archivo = load_workbook(ruta, read_only=True)
        hoja = archivo.active
        fila = hoja.max_row
        celdas = hoja["A1": "P" + str(fila)]

        # ######################################### #
        # ########## VALIDACION DEL EXEL ########## #
        # ######################################### #

        lista_vacia = []

        for i in celdas:
            datus = [celda.value for celda in i]
            lista_vacia.append(datus)

        lista_vacia.pop(0)
        # Limpiamos la base de datos
        for i in resultado:
            BaseDatos.delete().where(BaseDatos.orden == i).execute()

        for i in lista_vacia:
            # Tuve problemas al tratar de importar celdas vacias
            # Por eso se las verifican y si son nulas les cargo un string vacio
            if not i[2]:
                i[2] = " "
            if not i[3]:
                i[3] = " "
            if not i[4]:
                i[4] = " "
            if not i[5]:
                i[5] = " "
            if not i[6]:
                i[6] = " "
            if not i[7]:
                i[7] = " "
            if not i[8]:
                i[8] = " "
            if not i[9]:
                i[9] = " "
            if not i[10]:
                i[10] = " "
            if not i[11]:
                i[11] = " "
            if not i[12]:
                i[12] = " "
            if not i[13]:
                i[13] = " "
            if not i[14]:
                i[14] = " "
            if not i[15]:
                i[15] = " "
            if not i[1]:
                i[1] = " "
            # Guardamos los datos en la base de datos
            guard.guardar(
                i[2],
                i[3],
                i[4],
                i[5],
                i[6],
                i[7],
                i[8],
                i[9],
                i[10],
                i[11],
                i[12],
                i[13],
                i[14],
                i[15],
                i[1],
            )
        messagebox.showinfo(
            title="Exito",
            message="""
Base de datos importada desde el exel correctamente.
Cierre el programa para cargar la base de datos
            """,
        )


def exportacion(bango=True):
    fecha = datetime.now()
    nombre_final = "Base_datos-" + str(fecha.date()) + ".xlsx"
    archivo = Workbook()
    hoja = archivo.active
    resultado = BaseDatos.select()

    hoja["A1"] = "Orden"
    hoja["B1"] = "Fecha"
    hoja["C1"] = "Nombre"
    hoja["D1"] = "Apellido"
    hoja["E1"] = "Teléfono"
    hoja["F1"] = "Dirección"
    hoja["G1"] = "Tipo"
    hoja["H1"] = "Marca"
    hoja["I1"] = "Modelo"
    hoja["J1"] = "Falla"
    hoja["K1"] = "Otros"
    hoja["L1"] = "Estado"
    hoja["M1"] = "Costos"
    hoja["N1"] = "Total"
    hoja["O1"] = "Descripción"
    hoja["P1"] = "Notificación"

    co = 1

    for fila in resultado:
        co = co + 1
        hoja["A" + str(co)] = fila.orden
        hoja["B" + str(co)] = fila.fecha
        hoja["C" + str(co)] = fila.nombre
        hoja["D" + str(co)] = fila.apellido
        hoja["E" + str(co)] = fila.telefono
        hoja["F" + str(co)] = fila.direccion
        hoja["G" + str(co)] = fila.tipo
        hoja["H" + str(co)] = fila.marca
        hoja["I" + str(co)] = fila.modelo
        hoja["J" + str(co)] = fila.falla
        hoja["K" + str(co)] = fila.otros
        hoja["L" + str(co)] = fila.estado
        hoja["M" + str(co)] = fila.costo
        hoja["N" + str(co)] = fila.total
        hoja["O" + str(co)] = fila.descripcion
        hoja["P" + str(co)] = fila.notificacion
    archivo.save(nombre_final)
    if bango:
        messagebox.showinfo(
            title="Exito",
            message="Base de datos exportada a exel correctamente.",
        )
    return nombre_final
